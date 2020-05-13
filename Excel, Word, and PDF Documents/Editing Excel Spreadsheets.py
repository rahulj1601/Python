import openpyxl, os

wb = openpyxl.Workbook()

print(wb.sheetnames)

sheet = wb['Sheet'] #accesses a specific sheet
sheet['A1'] = 42
sheet['A2'] = 'Hello'

wb.save('example1.xlsx')

sheet2 = wb.create_sheet() #adding new sheet
sheet2.title = 'My New Sheet Name' #changing sheet name
print(wb.sheetnames)

wb.save('example2.xlsx')

wb.create_sheet(index=0, title='My Other Sheet')
# index makes this sheet the first sheet - changes the order of the sheets

wb.save('example3.xlsx')
