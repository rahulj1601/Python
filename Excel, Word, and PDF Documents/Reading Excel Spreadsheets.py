import openpyxl, os

os.chdir('/Users/rahul/Documents/Programming/Automating the Boring Stuff with Python /Excel, Word, and PDF Documents/')

workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))

sheet = workbook["Sheet1"]
print(type(sheet))

print(workbook.sheetnames) #returns sheet names

cell = sheet["A1"]
print(cell.value)

print(sheet["B1"].value)
print(sheet["C1"].value)

print(sheet.cell(row=1, column=2)) #gives the character-number reference of the cell

for i in range(1,8):
    print(i, sheet.cell(row=i, column=2).value)
